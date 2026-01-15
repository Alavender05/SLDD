import openpyxl
from openpyxl.styles import Font
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
import os
import numpy as np

# ==========================================
# 1. HELPER FUNCTIONS
# ==========================================

def get_value(wb, sheet_name, cell_ref):
    """Helper to get a value safely."""
    if sheet_name in wb.sheetnames:
        return wb[sheet_name][cell_ref].value
    return None

def clean_val(val):
    """Converts Excel value to float, handling None and strings."""
    if val is None or val == "":
        return 0.0
    if isinstance(val, (int, float, np.integer, np.floating)):
        return float(val)
    try:
        clean_str = str(val).replace("$", "").replace(",", "").replace("%", "").strip()
        return float(clean_str)
    except (ValueError, TypeError):
        return 0.0

def calc_growth(current, previous):
    """Calculates percentage growth between two values. Returns formatted string."""
    prev_float = clean_val(previous)
    curr_float = clean_val(current)
    
    if prev_float == 0:
        return "N/A"
    
    growth = (curr_float - prev_float) / prev_float
    return f"{growth:.2%}"

def style_output_sheet(ws):
    """Auto-adjust column widths and bold headers."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).startswith("---"):
                cell.font = Font(bold=True)
            if cell.value in ["Metric", "Income Range", "Description", "Year"]:
                cell.font = Font(bold=True)
                
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = (max_length + 2)

def add_conditional_formatting(ws, start_row, end_row, start_col_idx, end_col_idx):
    """Applies a 2-Color Scale (White -> Red) to the specified range."""
    start_col_char = get_column_letter(start_col_idx)
    end_col_char = get_column_letter(end_col_idx)
    cell_range = f"{start_col_char}{start_row}:{end_col_char}{end_row}"
    
    rule = ColorScaleRule(
        start_type='min', start_color='FFFFFF',
        end_type='max', end_color='FF0000'
    )
    ws.conditional_formatting.add(cell_range, rule)

# ==========================================
# 2. MAIN EXPORT FUNCTION
# ==========================================

def export_data_to_excel():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_filename = "TSP_305041135.xlsx"
    output_filename = "extracted_data.xlsx"

    input_path = os.path.join(script_dir, input_filename)
    output_path = os.path.join(script_dir, output_filename)

    print(f"Reading from: {input_filename}")
    print(f"Writing to:   {output_filename}\n")

    try:
        wb_source = openpyxl.load_workbook(input_path, data_only=True)
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = "Summary Report"

        # --- SECTION 1: T02 ---
        if "T02" in wb_source.sheetnames:
            sheet = wb_source["T02"]
            ws_out.append(["--- T02 MEDIAN / AVERAGE METRICS ---"])
            ws_out.append(["Metric", "2011", "2016", "2021"])
            
            metric_rows = [15, 17, 19, 21, 23]
            t02_metrics = []
            
            # Left side
            for r in metric_rows:
                name = sheet[f"A{r}"].value
                v11 = sheet[f"B{r}"].value
                v16 = sheet[f"C{r}"].value
                v21 = sheet[f"D{r}"].value
                if name: t02_metrics.append([name, v11, v16, v21])
            
            # Right side
            for r in metric_rows:
                name = sheet[f"F{r}"].value
                v11 = sheet[f"G{r}"].value
                v16 = sheet[f"H{r}"].value
                v21 = sheet[f"I{r}"].value
                if name: t02_metrics.append([name, v11, v16, v21])
                
            for row in t02_metrics: ws_out.append(row)

        ws_out.append([])
        ws_out.append([]) 

        # --- SECTION 2: SUMMARY TABLE (With Growth Rates) ---
        ws_out.append(["--- SUMMARY (2011 / 2016 / 2021) ---"])
        
        # UPDATED HEADERS
        ws_out.append(["Metric", "2011", "2016", "2021", "Growth '11-'16", "Growth '16-'21", "Total Growth '11-'21"])

        summary_items = [
            ("Total Persons Divorced",       ("T04", "L28"),  ("T04", "L48"),  ("T04", "L68")),
            ("Separate House",               ("T14a", "J13"), ("T14b", "J13"), ("T14c", "J13")),
            ("Flat or Apartment",            ("T14a", "J26"), ("T14b", "J26"), ("T14c", "J26")),
            ("Owned Outright",               ("T18", "G15"),  ("T18", "G34"),  ("T18", "G53")),
            ("Owned with a Mortgage",        ("T18", "G16"),  ("T18", "G35"),  ("T18", "G54")),
            ("Rented",                       ("T18", "G25"),  ("T18", "G44"),  ("T18", "G63")),
            ("Employed Worked Full Time",    ("T29", "D15"),  ("T29", "H15"),  ("T29", "L15")),
            ("Unemployment %",               ("T29", "D23"),  ("T29", "H23"),  ("T29", "L23")),
            ("Labour Force Participation",   ("T29", "D24"),  ("T29", "H24"),  ("T29", "L24")),
        ]

        for metric, (s11, c11), (s16, c16), (s21, c21) in summary_items:
            v11 = get_value(wb_source, s11, c11)
            v16 = get_value(wb_source, s16, c16)
            v21 = get_value(wb_source, s21, c21)
            
            # CALCULATE GROWTH RATES
            g_11_16 = calc_growth(v16, v11)
            g_16_21 = calc_growth(v21, v16)
            g_11_21 = calc_growth(v21, v11)

            ws_out.append([metric, v11, v16, v21, g_11_16, g_16_21, g_11_21])

        ws_out.append([])
        ws_out.append([]) 

        # --- SECTION 3: T24 MATRIX (With Heatmap) ---
        if "T24" in wb_source.sheetnames:
            sheet = wb_source["T24"]

            rent_labels = [
                "$1-$74", "$75-$99", "$100-$149", "$150-$199", "$200-$224",
                "$225-$274", "$275-$349", "$350-$449", "$450-$549",
                "$550-$649", "$650 or more"
            ]

            ws_out.append(["--- DATA FROM T24 (Income x Rent Matrix) ---"])
            ws_out.append(["Income Range"] + rent_labels)

            data_start_row = ws_out.max_row + 1 
            
            raw_rows = []
            for row in sheet.iter_rows(min_row=55, max_row=71, min_col=1, max_col=14, values_only=True):
                income_label = "" if row[0] is None else str(row[0]).strip()
                if income_label.upper().replace(" ", "") in ("2021CENSUS", "2021CENSUSYEAR", "CENSUSYEAR") or income_label == "":
                    continue

                rent_vals = row[1:12]  # B-L
                clean_row = [income_label] + [0 if v in (None, "") else v for v in rent_vals]
                raw_rows.append(clean_row)
                ws_out.append(clean_row)

            if raw_rows:
                numeric_matrix = [[clean_val(v) for v in r[1:]] for r in raw_rows]
                col_totals = np.sum(np.array(numeric_matrix, dtype=float), axis=0)
                totals_row = ["TOTAL"] + [int(x) if float(x).is_integer() else float(x) for x in col_totals]
                ws_out.append(totals_row)

            # Apply Heatmap
            data_end_row = data_start_row + len(raw_rows) - 1
            if len(raw_rows) > 0:
                print(f"Applying Heatmap to Rows {data_start_row}-{data_end_row}")
                add_conditional_formatting(ws_out, data_start_row, data_end_row, 2, 12)

            print("Extracted T24 data with Heatmap.")

        else:
            print("Warning: Sheet 'T24' not found.")

        style_output_sheet(ws_out)
        wb_out.save(output_path)
        print(f"\nSuccess! Open '{output_filename}' to see the report with growth rates.")

    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    export_data_to_excel()
