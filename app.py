import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np
import io
import requests
from bs4 import BeautifulSoup
import re


# ==========================================
# 1. SHARED EXCEL STYLES & HELPERS
# ==========================================

def get_header_style():
    return {
        "font": Font(bold=True, color="FFFFFF", size=11),
        "fill": PatternFill("solid", fgColor="4472C4"),
        "border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    }


def style_sheet_columns(ws):
    """Auto-adjust column widths."""
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column].width = (max_length + 2)


# ==========================================
# 2. TSP ANALYSIS LOGIC (FROM UPLOAD OR DOWNLOAD)
# ==========================================

def get_value(wb, sheet_name, cell_ref):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name][cell_ref].value
    return None


def clean_val(val):
    if val is None or val == "":
        return 0.0
    if isinstance(val, (int, float, np.integer, np.floating)):
        return float(val)
    try:
        clean_str = (
            str(val).replace("$", "").replace(",", "").replace("%", "").strip()
        )
        return float(clean_str)
    except (ValueError, TypeError):
        return 0.0


def calc_growth(current, previous):
    prev_float = clean_val(previous)
    curr_float = clean_val(current)
    if prev_float == 0:
        return "N/A"
    growth = (curr_float - prev_float) / prev_float
    return f"{growth:.2%}"


def get_midpoint(label):
    if not isinstance(label, str):
        return 0
    clean = label.replace("$", "").replace(",", "").strip()
    if "Negative" in clean or "Nil" in clean:
        return 0
    elif "or more" in clean:
        try:
            return float(clean.replace(" or more", "")) * 1.1
        except Exception:
            return 0
    elif "-" in clean:
        try:
            low, high = map(float, clean.split("-"))
            return (low + high) / 2
        except Exception:
            return 0
    return 0


def add_conditional_formatting(ws, start_row, end_row, start_col_idx, end_col_idx):
    start_col = get_column_letter(start_col_idx)
    end_col = get_column_letter(end_col_idx)
    cell_range = f"{start_col}{start_row}:{end_col}{end_row}"
    ws.conditional_formatting.add(
        cell_range,
        ColorScaleRule(
            start_type="min",
            start_color="FFFFFF",
            end_type="max",
            end_color="FF0000",
        ),
    )


def write_tsp_analysis_to_sheet(wb, uploaded_file, sheet_name="TSP Analysis"):
    """Reads uploaded file and adds a TSP Analysis sheet to the wb."""
    wb_source = openpyxl.load_workbook(uploaded_file, data_only=True)
    ws_out = wb.create_sheet(sheet_name)

    # --- T02 (per-area view only) ---
    if "T02" in wb_source.sheetnames:
        sheet = wb_source["T02"]
        ws_out.append(["--- T02 MEDIAN / AVERAGE METRICS ---"])
        ws_out.append(["Metric", "2011", "2016", "2021"])

        for cell in ws_out[2]:
            cell.font = Font(bold=True)

        rows = [15, 17, 19, 21, 23]
        for r in rows:
            if sheet[f"A{r}"].value:
                ws_out.append(
                    [
                        sheet[f"A{r}"].value,
                        sheet[f"B{r}"].value,
                        sheet[f"C{r}"].value,
                        sheet[f"D{r}"].value,
                    ]
                )
            if sheet[f"F{r}"].value:
                ws_out.append(
                    [
                        sheet[f"F{r}"].value,
                        sheet[f"G{r}"].value,
                        sheet[f"H{r}"].value,
                        sheet[f"I{r}"].value,
                    ]
                )
        ws_out.append([])
        ws_out.append([])

    # --- Summary Table ---
    ws_out.append(["--- SUMMARY (2011 / 2016 / 2021) ---"])
    header_row = [
        "Metric",
        "2011",
        "2016",
        "2021",
        "Growth '11-'16",
        "Growth '16-'21",
        "Total Growth '11-'21",
    ]
    ws_out.append(header_row)
    for cell in ws_out[ws_out.max_row]:
        cell.font = Font(bold=True)

    summary_items = [
        ("Total Persons Divorced", ("T04", "L28"), ("T04", "L48"), ("T04", "L68")),
        ("Separate House", ("T14a", "J13"), ("T14b", "J13"), ("T14c", "J13")),
        ("Flat or Apartment", ("T14a", "J26"), ("T14b", "J26"), ("T14c", "J26")),
        ("Owned Outright", ("T18", "G15"), ("T18", "G34"), ("T18", "G53")),
        ("Owned with a Mortgage", ("T18", "G16"), ("T18", "G35"), ("T18", "G54")),
        ("Rented", ("T18", "G25"), ("T18", "G44"), ("T18", "G63")),
        (
            "Employed Worked Full Time",
            ("T29", "D15"),
            ("T29", "H15"),
            ("T29", "L15"),
        ),
        ("Unemployment %", ("T29", "D23"), ("T29", "H23"), ("T29", "L23")),
        ("Labour Force Participation", ("T29", "D24"), ("T29", "H24"), ("T29", "L24")),
    ]
    for metric, s11, s16, s21 in summary_items:
        v11, v16, v21 = (
            get_value(wb_source, *s11),
            get_value(wb_source, *s16),
            get_value(wb_source, *s21),
        )
        v11_num = clean_val(v11)
        v16_num = clean_val(v16)
        v21_num = clean_val(v21)
        ws_out.append(
            [
                metric,
                v11_num,
                v16_num,
                v21_num,
                calc_growth(v16, v11),
                calc_growth(v21, v16),
                calc_growth(v21, v11),
            ]
        )
    ws_out.append([])
    ws_out.append([])

    # --- T24 Matrix ---
    if "T24" in wb_source.sheetnames:
        sheet = wb_source["T24"]
        ws_out.append(["--- DATA FROM T24 (Income x Rent Matrix) ---"])
        rent_labels = [
            "$1-$74",
            "$75-$99",
            "$100-$149",
            "$150-$199",
            "$200-$224",
            "$225-$274",
            "$275-$349",
            "$350-$449",
            "$450-$549",
            "$550-$649",
            "$650 or more",
        ]
        ws_out.append(["Income Range"] + rent_labels)
        for cell in ws_out[ws_out.max_row]:
            cell.font = Font(bold=True)

        start_row = ws_out.max_row + 1
        raw_rows = []
        for row in sheet.iter_rows(
            min_row=55, max_row=71, min_col=1, max_col=14, values_only=True
        ):
            lbl = "" if row[0] is None else str(row[0]).strip()
            if lbl and "CENSUS" not in lbl.upper():
                clean_row = [lbl] + [0 if v in (None, "") else v for v in row[1:12]]
                raw_rows.append(clean_row)
                ws_out.append(clean_row)

        if raw_rows:
            numeric_matrix = [[clean_val(v) for v in r[1:]] for r in raw_rows]
            totals = ["TOTAL"] + list(np.sum(numeric_matrix, axis=0))
            ws_out.append(totals)
            add_conditional_formatting(
                ws_out, start_row, start_row + len(raw_rows) - 1, 2, 12
            )

    style_sheet_columns(ws_out)


# ----- TSP summary + T24 extraction used for aggregation -----

def extract_tsp_summary(uploaded_file):
    """
    Extract summary metrics from TSP file (2011, 2016, 2021 values).
    Uses label lookups in T02/T04/T14/T18/T29 so it is robust to row shifts.
    """
    wb_source = openpyxl.load_workbook(uploaded_file, data_only=True)
    summary = {}

    def find_row_by_label(sheet, label, col="A"):
        label_lower = str(label).strip().lower()
        for row in range(1, sheet.max_row + 1):
            cell_val = sheet[f"{col}{row}"].value
            if cell_val is None:
                continue
            if str(cell_val).strip().lower() == label_lower:
                return row
        return None

    # T02 medians / averages
    if "T02" in wb_source.sheetnames:
        t02 = wb_source["T02"]
        t02_label_map = [
            ("Median age of persons", "Median age of persons", "left"),
            ("Median mortgage repayment ($/monthly)", "Median mortgage repayment ($/monthly)", "right"),
            ("Median total personal income ($/weekly)", "Median total personal income ($/weekly)", "left"),
            ("Median rent ($/weekly)(a)", "Median rent ($/weekly)(a)", "right"),
            ("Median total family income ($/weekly)", "Median total family income ($/weekly)", "left"),
            ("Average number of persons per bedroom", "Average number of persons per bedroom", "right"),
            ("Median total household income ($/weekly)", "Median total household income ($/weekly)", "left"),
            ("Average household size", "Average household size", "left"),
        ]
        for metric_name, label_text, side in t02_label_map:
            r = find_row_by_label(t02, label_text, col="A" if side == "left" else "F")
            if r is None:
                summary[metric_name] = [0.0, 0.0, 0.0]
                continue
            if side == "left":
                c11, c16, c21 = f"B{r}", f"C{r}", f"D{r}"
            else:
                c11, c16, c21 = f"G{r}", f"H{r}", f"I{r}"
            v11 = t02[c11].value
            v16 = t02[c16].value
            v21 = t02[c21].value
            summary[metric_name] = [clean_val(v11), clean_val(v16), clean_val(v21)]

    # T04 – Total Persons Divorced
    if "T04" in wb_source.sheetnames:
        t04 = wb_source["T04"]

        def get_t04(cell):
            return clean_val(t04[cell].value)

        summary["Total Persons Divorced"] = [
            get_t04("L28"),
            get_t04("L48"),
            get_t04("L68"),
        ]

    # T14a/b/c – Separate House, Flat or Apartment
    if all(s in wb_source.sheetnames for s in ["T14a", "T14b", "T14c"]):
        t14a, t14b, t14c = wb_source["T14a"], wb_source["T14b"], wb_source["T14c"]
        summary["Separate House"] = [
            clean_val(t14a["J13"].value),
            clean_val(t14b["J13"].value),
            clean_val(t14c["J13"].value),
        ]
        summary["Flat or Apartment"] = [
            clean_val(t14a["J26"].value),
            clean_val(t14b["J26"].value),
            clean_val(t14c["J26"].value),
        ]

    # T18 – tenure
    if "T18" in wb_source.sheetnames:
        t18 = wb_source["T18"]
        summary["Owned Outright"] = [
            clean_val(t18["G15"].value),
            clean_val(t18["G34"].value),
            clean_val(t18["G53"].value),
        ]
        summary["Owned with a Mortgage"] = [
            clean_val(t18["G16"].value),
            clean_val(t18["G35"].value),
            clean_val(t18["G54"].value),
        ]
        summary["Rented"] = [
            clean_val(t18["G25"].value),
            clean_val(t18["G44"].value),
            clean_val(t18["G63"].value),
        ]

    # T29 – labour force
    if "T29" in wb_source.sheetnames:
        t29 = wb_source["T29"]
        summary["Employed Worked Full Time"] = [
            clean_val(t29["D15"].value),
            clean_val(t29["H15"].value),
            clean_val(t29["L15"].value),
        ]
        summary["Unemployment %"] = [
            clean_val(t29["D23"].value),
            clean_val(t29["H23"].value),
            clean_val(t29["L23"].value),
        ]
        summary["Labour Force Participation"] = [
            clean_val(t29["D24"].value),
            clean_val(t29["H24"].value),
            clean_val(t29["L24"].value),
        ]

    return summary


def extract_tsp_t24_matrix(uploaded_file):
    """
    Extract T24 Income x Rent matrix from TSP.
    Returns list of rows: [label, col1_val, col2_val, ..., col11_val]
    """
    wb_source = openpyxl.load_workbook(uploaded_file, data_only=True)
    if "T24" not in wb_source.sheetnames:
        return []

    sheet = wb_source["T24"]
    rows = []
    for row in sheet.iter_rows(
        min_row=55, max_row=71, min_col=1, max_col=14, values_only=True
    ):
        lbl = "" if row[0] is None else str(row[0]).strip()
        if lbl and "CENSUS" not in lbl.upper():
            clean_row = [lbl] + [clean_val(v) for v in row[1:12]]
            rows.append(clean_row)

    return rows


# ----- Aggregation rules: which metrics average vs total -----

AVERAGE_METRICS = {
    "Median age of persons",
    "Median mortgage repayment ($/monthly)",
    "Median total personal income ($/weekly)",
    "Median rent ($/weekly)(a)",
    "Median total family income ($/weekly)",
    "Average number of persons per bedroom",
    "Median total household income ($/weekly)",
    "Average household size",
    "Unemployment %",
    "Labour Force Participation",
}

TOTAL_METRICS = {
    "Total Persons Divorced",
    "Separate House",
    "Flat or Apartment",
    "Owned Outright",
    "Owned with a Mortgage",
    "Rented",
    "Employed Worked Full Time",
}


def aggregate_tsp_summaries(all_summaries):
    """
    all_summaries: list of dicts from extract_tsp_summary()
    Returns dict metric -> [2011, 2016, 2021].
    """
    if not all_summaries:
        return {}

    agg = {}
    counts = {}

    for summary in all_summaries:
        for metric_name, vals in summary.items():
            if metric_name not in agg:
                agg[metric_name] = np.array([0.0, 0.0, 0.0], dtype=float)
                counts[metric_name] = 0
            agg[metric_name] += np.array(vals, dtype=float)
            counts[metric_name] += 1

    out = {}
    for metric, totals in agg.items():
        n = counts[metric]
        if metric in AVERAGE_METRICS:
            out[metric] = list(totals / max(n, 1))
        else:
            out[metric] = list(totals)

    return out


def aggregate_t24_matrices(all_t24_matrices):
    """
    Sum Income x Rent matrices cell-wise across areas.
    """
    if not all_t24_matrices:
        return []

    grouped = {}
    for matrix in all_t24_matrices:
        for row in matrix:
            label = row[0]
            values = row[1:]
            if label not in grouped:
                grouped[label] = np.array(values, dtype=float)
            else:
                grouped[label] += np.array(values, dtype=float)

    result = []
    for label, totals in grouped.items():
        result.append([label] + list(totals))

    return result


def build_aggregated_tsp_sheet(wb, aggregated_summary, aggregated_t24):
    """
    Create aggregated TSP sheet with averaged/totals metrics and T24 matrix.
    """
    ws = wb.create_sheet("Aggregated TSP")

    ws.append(["--- AGGREGATED TSP SUMMARY (Averages & Totals across areas) ---"])
    ws.append(["Metric", "2011", "2016", "2021"])
    for cell in ws[2]:
        cell.font = Font(bold=True)

    for metric_name, values in aggregated_summary.items():
        ws.append([metric_name, values[0], values[1], values[2]])

    ws.append([])
    ws.append([])

    ws.append(["--- AGGREGATED T24 MATRIX (Totals across areas) ---"])
    rent_labels = [
        "$1-$74",
        "$75-$99",
        "$100-$149",
        "$150-$199",
        "$200-$224",
        "$225-$274",
        "$275-$349",
        "$350-$449",
        "$450-$549",
        "$550-$649",
        "$650 or more",
    ]
    ws.append(["Income Range"] + rent_labels)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    start_row = ws.max_row + 1
    for row in aggregated_t24:
        ws.append(row)

    if aggregated_t24:
        add_conditional_formatting(
            ws, start_row, start_row + len(aggregated_t24) - 1, 2, 12
        )

    style_sheet_columns(ws)


# ==========================================
# 3. SCRAPING LOGIC (ONLINE DATA)
# ==========================================

BASE_URLS = {
    2011: "https://www.abs.gov.au/census/find-census-data/quickstats/2011/{}",
    2016: "https://www.abs.gov.au/census/find-census-data/quickstats/2016/{}",
    2021: "https://www.abs.gov.au/census/find-census-data/quickstats/2021/{}",
}

METRICS = [
    {
        "name": "Average number of people per household",
        "unit": "",
        "variants": [
            "Average number of people per household",
            "Average people per household",
        ],
    },
    {
        "name": "Median weekly household income",
        "unit": "$",
        "variants": ["Median weekly household income"],
    },
    {
        "name": "Less than $650 total household weekly income (a)",
        "unit": "%",
        "variants": ["Less than $650 total household weekly income (a)"],
    },
    {
        "name": "More than $3,000 total household weekly income (a)",
        "unit": "%",
        "variants": ["More than $3,000 total household weekly income (a)"],
    },
    {
        "name": "Median monthly mortgage repayments",
        "unit": "$",
        "variants": ["Median monthly mortgage repayments"],
    },
    {
        "name": "Owned outright",
        "unit": "%",
        "variants": ["Owned outright", "Owned Outright"],
    },
    {
        "name": "Owned with a mortgage",
        "unit": "%",
        "variants": ["Owned with a mortgage", "Owned with a Mortgage"],
    },
    {"name": "Rented", "unit": "%", "variants": ["Rented"]},
    {
        "name": "Median weekly rent",
        "unit": "$",
        "variants": [
            "Median weekly rent",
            "Median weekly rent (a)",
            "Median weekly rent (b)",
        ],
    },
    {
        "name": "Rent payments less than 30% of household income",
        "unit": "%",
        "variants": [
            "Renter households where rent payments are less than or equal to 30% of household income (b)",
            "Households where rent payments are less than 30% of household income",
        ],
    },
    {
        "name": "Rent payments 30% or more of household income",
        "unit": "%",
        "variants": [
            "Renter households with rent payments greater than 30% of household income (b)",
            "Households where rent payments are 30%, or greater, of household income",
            "Households with rent payments greater than or equal to 30% of household income",
        ],
    },
    {
        "name": "Mortgage payments less than 30% of household income",
        "unit": "%",
        "variants": [
            "Owner with mortgage households where mortgage repayments are less than or equal to 30% of household income (a)",
            "Households where mortgage payments are less than 30% of household income",
            "Households where mortgage repayments are less than 30% of household income",
        ],
    },
    {
        "name": "Mortgage payments 30% or more of household income",
        "unit": "%",
        "variants": [
            "Owner with mortgage households with mortgage repayments greater than 30% of household income (a)",
            "Households where mortgage payments are 30%, or greater, of household income",
        ],
    },
    {"name": "Worked full-time", "unit": "%", "variants": ["Worked full-time"]},
    {"name": "Separate house", "unit": "%", "variants": ["Separate house"]},
    {
        "name": "Flat, unit or apartment",
        "unit": "%",
        "variants": ["Flat or apartment", "Flat, unit or apartment"],
    },
]


def get_quickstats_tables(area_code, year):
    url = BASE_URLS[year].format(area_code)
    try:
        resp = requests.get(url, timeout=15)
        resp.raise_for_status()
    except Exception:
        return None, None, None
    soup = BeautifulSoup(resp.text, "html.parser")
    tables = soup.find_all("table")
    h1 = soup.find("h1")
    area_name = h1.get_text(strip=True) if h1 else f"Area {area_code}"
    return tables, area_name, url


def extract_metric_value(tables, variants):
    if not tables:
        return None
    lower_variants = [v.lower() for v in variants]
    for table in tables:
        for tr in table.find_all("tr"):
            cells = [c.get_text(strip=True) for c in tr.find_all(["th", "td"])]
            if cells:
                label = cells[0].strip().lower()
                if any(v in label for v in lower_variants):
                    return (
                        cells[2].strip()
                        if len(cells) > 2
                        else (cells[1].strip() if len(cells) > 1 else None)
                    )
    return None


def extract_all_metrics(area_code, year):
    tables, area_name, url = get_quickstats_tables(area_code, year)
    if tables is None:
        return None
    
    result = {"area_code": area_code, "area_name": area_name, "year": year, "url": url}
    
    # NEW: Extract total "People" count (first data row of first table)
    people_count = None
    if tables:
        first_table = tables[0]
        for tr in first_table.find_all("tr"):
            cells = [c.get_text(strip=True) for c in tr.find_all(["th", "td"])]
            if len(cells) >= 2 and "people" in cells[0].lower():
                # Extract number from cells[1] (handles commas, etc.)
                raw_people = cells[1].replace(",", "").strip()
                try:
                    people_count = int(float(raw_people))
                    break
                except (ValueError, TypeError):
                    pass
    result["People"] = people_count or 0
    
    # Your existing metrics (unchanged)
    for m in METRICS:
        result[m["name"]] = extract_metric_value(tables, m["variants"])
    
    return result



def write_scraped_data_to_sheet(wb, data_dict, sheet_name="Online QuickStats"):
    """Adds an Online QuickStats sheet to the wb with proper % formatting."""
    ws = wb.create_sheet(sheet_name)

    styles = get_header_style()

    headers = ["Metric", "Unit", "", "2011", "2016", "2021"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = styles["fill"]
        cell.font = styles["font"]
        cell.alignment = styles["alignment"]
        cell.border = styles["border"]

    latest_data = data_dict.get(2021, data_dict.get(2016, data_dict.get(2011)))
    if latest_data:
        ws.append(["Area Code", "", "", latest_data["area_code"], "", ""])
        ws.append(["Area Name", "", "", latest_data["area_name"], "", ""])
        ws.append(["Source URL", "", "", latest_data["url"], "", ""])
        ws.append([])

    for m in METRICS:
        row = [m["name"], m["unit"], ""]
        for year in [2011, 2016, 2021]:
            val = data_dict.get(year, {}).get(m["name"], "—")
            row.append(val if val else "—")
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        unit = row[1].value
        for cell in row[3:6]:
            cell.alignment = Alignment(horizontal="right")
        for cell in row[3:6]:
            val = cell.value
            if val in (None, "—", ""):
                continue
            try:
                clean_str = (
                    str(val).replace("%", "").replace(",", "").replace("$", "").strip()
                )
                num = float(clean_str)
                if unit == "%":
                    cell.value = num / 100.0
                    cell.number_format = "0.0%"
                else:
                    cell.value = num
            except (ValueError, TypeError):
                pass

    ws.column_dimensions["A"].width = 50
    for col in ["D", "E", "F"]:
        ws.column_dimensions[col].width = 15

def build_2021_summary_workbook(multi_quickstats_dicts):
    """
    Create 2021 summary with People column + weighted averages by population.
    """
    wb_2021 = openpyxl.Workbook()
    ws = wb_2021.active
    ws.title = "2021 Demographics"

    styles = get_header_style()

    # Extract all 2021 entries with People counts
    area_entries = []
    for d in multi_quickstats_dicts:
        entry = d.get(2021)
        if entry and entry.get("People") is not None:
            area_entries.append(entry)

    if not area_entries:
        return None

    # NEW HEADER: People | Demographics | Area1 | Area2 | ... | Weighted Avg
    headers = ["People", "Demographics"]
    for e in area_entries:
        label = f"{e['area_code']} - {e['area_name'][:30]}..." if len(e['area_name']) > 30 else e['area_name']
        headers.append(label)
    headers.append("Weighted Avg")

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = styles["fill"]
        cell.font = styles["font"]
        cell.alignment = styles["alignment"]
        cell.border = styles["border"]

    # ROW 2: Show actual People counts (for verification)
    people_row = ["Total People"] + [""]  # People | "" (demographic name blank)
    all_populations = []
    for e in area_entries:
        people = e.get("People", 0)
        people_row.append(people)
        all_populations.append(people)
    people_row.append(f"={chr(66)}3")  # Formula: =SUM(C3:Z3) for total pop
    ws.append(people_row)
    
    total_population_cell = f"{get_column_letter(len(headers))}{ws.max_row}"
    ws[total_population_cell].font = Font(bold=True)

    # Process each metric
    for metric_idx, m in enumerate(METRICS, start=1):
        name = m["name"]
        unit = m["unit"]
        
        row_vals = [""]  # Column A: blank for metrics (People already shown above)
        row_vals.append(name)  # Column B: metric name
        
        # Extract values and populations for this metric
        metric_values = []
        area_populations = []
        
        for e in area_entries:
            raw_val = e.get(name)
            if raw_val in (None, "—", ""):
                metric_values.append(0)
                area_populations.append(0)
                row_vals.append("")
                continue
            
            try:
                # Clean and convert
                clean_str = str(raw_val).replace("%", "").replace(",", "").replace("$", "").strip()
                num_val = float(clean_str)
                
                # Store raw numeric value
                metric_values.append(num_val)
                area_populations.append(e.get("People", 0) or 0)
                
                # Display value (format % properly)
                display_val = num_val / 100.0 if unit == "%" else num_val
                row_vals.append(display_val)
                
            except (ValueError, TypeError):
                row_vals.append("")
                metric_values.append(0)
                area_populations.append(0)

        # CALCULATE WEIGHTED AVERAGE
        total_pop = sum(all_populations)  # Total population across ALL areas
        weighted_sum = 0.0
        valid_weights = 0
        
        for val, pop in zip(metric_values, area_populations):
            if pop > 0:
                weight = pop / total_pop  # population_i / total_population
                weighted_sum += val * weight
                valid_weights += 1
        
        if total_pop > 0 and valid_weights > 0:
            weighted_avg_raw = weighted_sum
            
            # Format for display
            if unit == "%":
                weighted_avg_display = weighted_avg_raw / 100.0
                row_vals.append(weighted_avg_display)
            else:
                row_vals.append(weighted_avg_raw)
        else:
            row_vals.append("N/A")

        ws.append(row_vals)

    # Auto-format columns
    ws.column_dimensions["A"].width = 12      # People column
    ws.column_dimensions["B"].width = 50      # Metric names
    ws.column_dimensions["C"].width = 12      # Total pop formula
    
    for col in range(4, ws.max_column):  # Area columns
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Format numbers
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):  # Metrics rows only
        metric_name = row[1].value
        unit = next((m["unit"] for m in METRICS if m["name"] == metric_name), "")
        
        for cell_idx in range(3, len(row)-1):  # Area columns only
            cell = row[cell_idx]
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
                if unit == "%":
                    cell.number_format = "0.0%"
        
        # Weighted Avg column
        avg_cell = row[-1]
        if isinstance(avg_cell.value, (int, float)):
            avg_cell.alignment = Alignment(horizontal="right")
            avg_cell.font = Font(bold=True)
            if unit == "%":
                avg_cell.number_format = "0.0%"

    return wb_2021

# ==========================================
# 4. ABS TIME SERIES PROFILE DOWNLOAD
# ==========================================

def download_tsp_for_area(area_code: str, year: int = 2021):
    """
    Download the Time Series Profile XLSX for the given area code and year.
    Returns a BytesIO object or None.
    """
    area_code = area_code.strip()
    base = "https://www.abs.gov.au"
    tsp_path = (
        f"/census/find-census-data/community-profiles/{year}/{area_code}/download/"
        f"TSP_{area_code}.xlsx"
    )
    tsp_url = base + tsp_path

    try:
        resp = requests.get(tsp_url, timeout=30)
        resp.raise_for_status()
    except Exception as e:
        st.warning(f"Could not download TSP file from {tsp_url}: {e}")
        return None

    return io.BytesIO(resp.content)


# ==========================================
# 5. STREAMLIT APP LOGIC
# ==========================================

def main():
    st.set_page_config(page_title="Census Data Tool", layout="wide", page_icon="📊")
    st.title("📊 Australian Census Data Combiner")
    st.write("Combine TSP analysis and Online QuickStats into integrated Excel reports.")
    st.markdown(
        "1. **Analyzes a Time Series Profile (TSP)** – Upload or auto-download from ABS"
    )
    st.markdown(
        "2. **Scrapes Online ABS QuickStats** – Fetches summary stats for a given Area Code"
    )
    st.markdown(
        "3. **Multi-area aggregation** – Auto-downloads TSP files and aggregates metrics across areas."
    )

    col1, col2 = st.columns(2)

    with col1:
        st.subheader("1. Time Series Profile (TSP)")
        uploaded_file = st.file_uploader(
            "Upload TSP_*.xlsx file",
            type=["xlsx"],
            help="Skip if you want to auto-download instead.",
        )

    with col2:
        st.subheader("2. Area Code & Options")
        area_code_input = st.text_input(
            "Enter ABS Area Code (e.g. 3GBRI):",
            help="Required for QuickStats scraping. Leave empty if only uploading TSP file.",
        ).strip().upper()

        auto_fetch_tsp = st.checkbox(
            "Auto-download Time Series Profile from ABS",
            value=False,
            help="If checked, the app will fetch the TSP XLSX from the ABS Community Profile instead of requiring an upload.",
        )

    st.markdown("---")
    st.subheader("3. Multi-area aggregation (optional)")
    st.write(
        "**Auto-downloads TSP files, averages relevant metrics, totals counts, and totals Income × Rent matrices.**"
    )

    multi_area_codes_raw = st.text_area(
        "Enter multiple ABS Area Codes (comma or newline separated):",
        help="The app will auto-download TSP files and aggregate metrics across all areas.",
    )

    multi_area_codes = [
        c.strip().upper()
        for c in re.split(r"[,\n]+", multi_area_codes_raw)
        if c.strip()
    ]

    create_separate_2021_file = st.checkbox(
        "Generate separate 2021 Demographics Excel file (multi-area)",
        value=False,
        help="Creates an Excel with 2021 metrics across all multi-area codes.",
    )

    if st.button("Generate Reports", type="primary"):
        if not uploaded_file and not (area_code_input and auto_fetch_tsp) and not (
            multi_area_codes
        ):
            st.error(
                "Please provide: (1) A TSP file upload OR (2) Area Code with auto-download OR (3) multi-area codes."
            )
            return

        wb_out = openpyxl.Workbook()
        if "Sheet" in wb_out.sheetnames:
            del wb_out["Sheet"]

        tsp_workbook_source = None

        # Single TSP (do not block QuickStats if download fails)
        if uploaded_file:
            tsp_workbook_source = uploaded_file
            st.success("✅ TSP file provided (upload)")
        elif area_code_input and auto_fetch_tsp:
            with st.spinner(f"Downloading Time Series Profile for {area_code_input}..."):
                tsp_bytes = download_tsp_for_area(area_code_input)
                if tsp_bytes is None:
                    st.warning(
                        f"Could not download Time Series Profile for {area_code_input}. "
                        "Skipping TSP analysis but QuickStats scraping will still run."
                    )
                    tsp_workbook_source = None
                else:
                    tsp_workbook_source = tsp_bytes
                    st.success(f"✅ Time Series Profile downloaded for {area_code_input}")

        if tsp_workbook_source:
            with st.spinner("Processing TSP workbook..."):
                try:
                    write_tsp_analysis_to_sheet(
                        wb_out, tsp_workbook_source, "TSP Analysis"
                    )
                    st.success("✅ TSP Analysis sheet created")
                except Exception as e:
                    st.warning(
                        f"Error processing TSP workbook. Skipping TSP analysis but continuing with QuickStats. Details: {e}"
                    )

        # Single QuickStats (always allowed to run if area_code_input is set)
        data_by_year_single = None
        if area_code_input:
            with st.spinner(f"Scraping QuickStats for {area_code_input}..."):
                data_by_year = {}
                years = [2011, 2016, 2021]
                progress_bar = st.progress(0)

                valid_data = False
                for i, year in enumerate(years):
                    data = extract_all_metrics(area_code_input, year)
                    if data:
                        data_by_year[year] = data
                        valid_data = True
                    progress_bar.progress((i + 1) / len(years))

                if valid_data:
                    data_by_year_single = data_by_year
                    write_scraped_data_to_sheet(
                        wb_out, data_by_year, sheet_name="Online QuickStats"
                    )
                    st.success(f"✅ QuickStats sheet created for {area_code_input}")
                else:
                    st.warning(f"Could not find QuickStats data for {area_code_input}")

        # Multi-area TSP + QuickStats
        multi_quickstats_dicts = []
        if data_by_year_single:
            multi_quickstats_dicts.append(data_by_year_single)

        all_tsp_summaries = []
        all_t24_matrices = []

        for code in multi_area_codes:
            if code == area_code_input:
                continue

            # TSP download is optional – failures do not block QuickStats
            with st.spinner(f"Downloading and processing TSP for {code}..."):
                tsp_bytes = download_tsp_for_area(code)
                if tsp_bytes is None:
                    st.warning(
                        f"Could not download TSP for {code}. Skipping TSP but still running QuickStats."
                    )
                else:
                    try:
                        summary = extract_tsp_summary(tsp_bytes)
                        t24 = extract_tsp_t24_matrix(tsp_bytes)
                        all_tsp_summaries.append(summary)
                        all_t24_matrices.append(t24)
                        st.success(f"✅ TSP data extracted for {code}")
                    except Exception as e:
                        st.warning(
                            f"Error processing TSP for {code}. Skipping TSP but continuing with QuickStats. Details: {e}"
                        )

            with st.spinner(f"Scraping QuickStats for {code}..."):
                data_by_year_extra = {}
                valid_extra = False
                for year in [2011, 2016, 2021]:
                    d = extract_all_metrics(code, year)
                    if d:
                        data_by_year_extra[year] = d
                        valid_extra = True
                if valid_extra:
                    multi_quickstats_dicts.append(data_by_year_extra)
                    st.success(f"✅ QuickStats scraped for {code}")
                else:
                    st.warning(f"No QuickStats data for {code}")

        # Aggregated TSP sheet (only if at least one TSP was available)
        if all_tsp_summaries:
            agg_summary = aggregate_tsp_summaries(all_tsp_summaries)
            agg_t24 = aggregate_t24_matrices(all_t24_matrices)
            build_aggregated_tsp_sheet(wb_out, agg_summary, agg_t24)
            st.success(
                f"✅ Aggregated TSP sheet created (across {len(all_tsp_summaries)} areas)"
            )

        # Save main workbook
        if len(wb_out.sheetnames) > 0:
            buffer = io.BytesIO()
            wb_out.save(buffer)
            buffer.seek(0)

            if area_code_input:
                file_label = f"{area_code_input}_Combined_Report.xlsx"
            else:
                file_label = "Census_Combined_Report.xlsx"

            st.download_button(
                label="📥 Download Main Excel Report",
                data=buffer,
                file_name=file_label,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            st.info(
                f"Main report contains {len(wb_out.sheetnames)} sheet(s): "
                f"{', '.join(wb_out.sheetnames)}"
            )
        else:
            st.error("No data was generated in the main report. Please check your inputs.")

        # Separate 2021 workbook
        if create_separate_2021_file and multi_quickstats_dicts:
            wb_2021 = build_2021_summary_workbook(multi_quickstats_dicts)
            if wb_2021 is None:
                st.warning(
                    "Could not create 2021 Demographics workbook (no 2021 data found)."
                )
            else:
                buf_2021 = io.BytesIO()
                wb_2021.save(buf_2021)
                buf_2021.seek(0)
                st.download_button(
                    label="📋 Download 2021 Demographics Summary",
                    data=buf_2021,
                    file_name="2021_Demographics_Summary.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
        elif create_separate_2021_file:
            st.warning(
                "To generate the 2021 Demographics file, provide multi-area codes or QuickStats data."
            )


if __name__ == "__main__":
    main()